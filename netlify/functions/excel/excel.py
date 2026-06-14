import json
import base64
import io
import zipfile
import os
import re
from datetime import datetime, timedelta

try:
    import openpyxl
    from openpyxl.utils import coordinate_to_tuple
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# VERSAO: 3
TEMPLATES_DIR = os.path.join(os.path.dirname(__file__), '..', '..', '..', 'templates')

MESES_PT = ['janeiro','fevereiro','março','abril','maio','junho',
            'julho','agosto','setembro','outubro','novembro','dezembro']

VARS = {
    'JB':  [0.21,0.41,0.61,0.81,0.61,0.11,0.21,0.41,0.61,0.81,0.61,0.11,
            0.21,0.41,0.61,0.81,0.61,0.11,0.21,0.41,0.61,0.81,0.61,0.11,
            0.21,0.41,0.61,0.81,0.61,0.11,0.21,0.41],
    'LC4': [0.97,0.087,0.077,0.067,0.57,0.17,0.097,0.087,0.077,0.067,0.057,0.17,
            0.097,0.087,0.077,0.067,0.057,0.17,0.097,0.087,0.077,0.067,0.057,0.17,
            0.097,0.087,0.077,0.067,0.057,0.17,0.097,0.087],
    'MF':  [0.12,0.032,0.52,0.082,0.012,0.012,0.032,0.52,0.082,0.012,0.012,0.032,
            0.52,0.082,0.012,0.012,0.032,0.52,0.082,0.012,0.012,0.032,0.52,0.082,
            0.012,0.012,0.032,0.52,0.082,0.012,0.012,0.032],
    'RPL': [0.32,0.62,0.92,0.12,0.32,0.62,0.92,0.12,0.32,0.62,0.92,0.12,
            0.32,0.62,0.92,0.12,0.32,0.62,0.92,0.12,0.32,0.62,0.92,0.12,
            0.32,0.62,0.92,0.12,0.32,0.62,0.92,0.12],
    'WB':  [0.12,0.32,0.062,0.092,0.162,0.12,0.032,0.062,0.92,0.162,0.012,0.032,
            0.062,0.092,0.162,0.012,0.032,0.062,0.092,0.162,0.012,0.032,0.062,0.092,
            0.162,0.012,0.032,0.062,0.092,0.162,0.012,0.032],
    'WR':  [0.67,0.067,0.087,0.097,0.13,0.17,0.067,0.087,0.097,0.13,0.067,0.067,
            0.087,0.097,0.13,0.067,0.067,0.087,0.097,0.13,0.067,0.067,0.087,0.097,
            0.13,0.067,0.067,0.087,0.097,0.13,0.067,0.067],
}

SHEET_MAP = {'JB':'JB','LC4':'L.C 4R','MF':'MF','RPL':'RPL','WB':'WB','WR':'WR'}

MIME_MAP = {
    'png':'image/png','jpg':'image/jpeg','jpeg':'image/jpeg',
    'gif':'image/gif','bmp':'image/bmp','emf':'image/x-emf','wmf':'image/x-wmf',
}

def set_cell(ws, cell_ref, value):
    row, col = coordinate_to_tuple(cell_ref)
    ws.cell(row=row, column=col).value = value

def parse_float(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = re.sub(r'[^\d,.]', '', str(value).strip())
    if not s:
        return 0.0
    if ',' in s and '.' in s:
        if s.rfind(',') > s.rfind('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif ',' in s:
        s = s.replace(',', '.')
    parts = s.split('.')
    if len(parts) > 2:
        s = ''.join(parts[:-1]) + '.' + parts[-1]
    try:
        return float(s)
    except ValueError:
        return 0.0

def _fix_targets(xml, prefix_map):
    def replacer(m):
        v = m.group(1)
        for a, r in prefix_map.items():
            if v.startswith(a):
                v = r + v[len(a):]
                break
        return f'Target="{v}"'
    return re.sub(r'Target="([^"]+)"', replacer, xml)

def reinjetar_imagens(xlsx_bytes, tmpl_path):
    nf = {}
    with zipfile.ZipFile(io.BytesIO(xlsx_bytes)) as z:
        for n in z.namelist():
            nf[n] = z.read(n)

    res = {}
    with zipfile.ZipFile(tmpl_path) as z:
        for n in z.namelist():
            res[n] = z.read(n)

    DATA = ('xl/worksheets/', 'xl/styles.xml', 'xl/workbook.xml', 'docProps/')
    for n, d in nf.items():
        if '_rels' in n:
            continue
        if any(n.startswith(p) or n == p for p in DATA):
            res[n] = d

    wbr = 'xl/_rels/workbook.xml.rels'
    if wbr in nf:
        res[wbr] = _fix_targets(nf[wbr].decode(), {
            '/xl/worksheets/': 'worksheets/',
            '/xl/styles':      'styles',
            '/xl/theme':       'theme',
            '/xl/sharedStrings': 'sharedStrings',
        }).encode()

    sr = 'xl/worksheets/_rels/sheet1.xml.rels'
    if sr in nf:
        res[sr] = _fix_targets(nf[sr].decode(), {
            '/xl/drawings/':        '../drawings/',
            '/xl/printerSettings/': '../printerSettings/',
        }).encode()

    ctk = '[Content_Types].xml'
    if ctk in nf:
        ct = nf[ctk].decode()
        for n in res:
            if 'xl/media/' in n:
                ext = n.rsplit('.', 1)[-1].lower()
                mime = MIME_MAP.get(ext, f'image/{ext}')
                if f'Extension="{ext}"' not in ct:
                    ct = ct.replace('<Override ', f'<Default Extension="{ext}" ContentType="{mime}"/><Override ', 1)
        res[ctk] = ct.encode()

    rf = res.get(wbr, b'').decode()
    cf = res.get(ctk, b'').decode()
    if 'sharedStrings' not in rf and 'sharedStrings' not in cf:
        res.pop('xl/sharedStrings.xml', None)

    res.pop('xl/calcChain.xml', None)

    if 'xl/workbook.xml' in res:
        w = res['xl/workbook.xml'].decode()
        if '#REF!' in w:
            w = re.sub(r'<definedNames>.*?</definedNames>', '', w, flags=re.DOTALL)
        res['xl/workbook.xml'] = w.replace('<definedNames/>', '').encode()

    out = io.BytesIO()
    with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as z:
        for n, d in res.items():
            z.writestr(n, d)
    return out.getvalue()

def preencher_egf(itens, cli, odata, onum, tipo):
    path = os.path.join(TEMPLATES_DIR, 'Engefal.xlsx')
    wb   = openpyxl.load_workbook(path)
    ws   = wb['EGF']

    set_cell(ws, 'I3',  odata)
    set_cell(ws, 'O3',  odata)
    set_cell(ws, 'O4',  odata - timedelta(days=7))
    set_cell(ws, 'I5',  int(onum) if onum else 1)
    set_cell(ws, 'H15', tipo)

    if cli:
        set_cell(ws, 'D9',  cli.get('nome', ''))
        set_cell(ws, 'D10', cli.get('nome2', ''))
        set_cell(ws, 'G9',  cli.get('contato', ''))
        set_cell(ws, 'G10', cli.get('cnpj', ''))
        set_cell(ws, 'I9',  cli.get('end', ''))
        set_cell(ws, 'I10', cli.get('tel', ''))

    total = 0.0
    for i in range(35):
        row = 18 + i
        it  = itens[i] if i < len(itens) else None
        if it:
            vu  = round(parse_float(it.get('valorUnitario', 0)), 2)
            qtd = parse_float(it.get('qtd', 1))
            vt  = round(vu * qtd, 2)
            total += vt
            set_cell(ws, f'E{row}', qtd)
            set_cell(ws, f'F{row}', it.get('unidade', 'Unid'))
            set_cell(ws, f'G{row}', it.get('produto') or it.get('descricao', ''))
            set_cell(ws, f'H{row}', vu)
            set_cell(ws, f'I{row}', vt)
        else:
            set_cell(ws, f'E{row}', None)
            set_cell(ws, f'F{row}', None)
            set_cell(ws, f'G{row}', None)
            set_cell(ws, f'H{row}', 0)
            set_cell(ws, f'I{row}', 0)

    set_cell(ws, 'I53', round(total, 2))
    buf = io.BytesIO()
    wb.save(buf)
    return reinjetar_imagens(buf.getvalue(), path)

def preencher_concorrente(cid, itens, cli, data_conc, tipo, vars_ia=None):
    path = os.path.join(TEMPLATES_DIR, f'{cid}.xlsx')
    wb   = openpyxl.load_workbook(path)
    ws   = wb[SHEET_MAP[cid]]

    nome    = cli.get('nome', '')    if cli else ''
    end     = cli.get('end', '')     if cli else ''
    cnpj    = cli.get('cnpj', '')    if cli else ''
    contato = cli.get('contato', '') if cli else ''
    tel     = cli.get('tel', '')     if cli else ''
    escola  = cli.get('nome2', '')   if cli else ''
    d       = data_conc
    slash   = f"{d.day:02d}/{d.month:02d}/{d.year}"
    mes     = MESES_PT[d.month - 1]

    if cid == 'JB':
        set_cell(ws, 'C17', f'Cliente:  {nome}')
        set_cell(ws, 'C18', f'Endereço: {end}')
        set_cell(ws, 'C19', f'Contato: {contato}')
        set_cell(ws, 'C21', f'Segue abaixo nosso orçamento de: {tipo}')
        set_cell(ws, 'C66', f'São Paulo, {d.day} de {mes} {d.year}')
        start_row, dc, qc, vc, tc, tot = 24,'C','D','E','G','G56'
    elif cid == 'LC4':
        set_cell(ws, 'B5',  nome)
        set_cell(ws, 'B6',  end)
        set_cell(ws, 'B7',  tipo)
        set_cell(ws, 'B51', f'São Paulo, {slash}')
        start_row, dc, qc, vc, tc, tot = 10,'B','C','D','E','E46'
    elif cid == 'MF':
        set_cell(ws, 'C5',  nome)
        set_cell(ws, 'B6',  end)
        set_cell(ws, 'B8',  tipo.upper())
        set_cell(ws, 'B53', f'São Paulo, {slash}')
        start_row, dc, qc, vc, tc, tot = 10,'C','D','E','F','F45'
    elif cid == 'RPL':
        set_cell(ws, 'C12', d)
        set_cell(ws, 'C13', cnpj)
        set_cell(ws, 'C14', escola)
        set_cell(ws, 'C15', end)
        set_cell(ws, 'C16', f'Orçamento de:  {tipo}')
        set_cell(ws, 'C58', f'{d.day} de {mes} de {d.year}')
        start_row, dc, qc, vc, tc, tot = 19,'C','D','E','F','F51'
    elif cid == 'WB':
        set_cell(ws, 'B3',  nome)
        set_cell(ws, 'B4',  end)
        set_cell(ws, 'B6',  tipo)
        set_cell(ws, 'E47', d)
        start_row, dc, qc, vc, tc, tot = 8,'B','C','D','E','E44'
    elif cid == 'WR':
        set_cell(ws, 'C14', nome)
        set_cell(ws, 'C15', end)
        set_cell(ws, 'C16', contato)
        set_cell(ws, 'C17', tel)
        set_cell(ws, 'B22', tipo)
        set_cell(ws, 'B60', f'São Paulo, {d.day} {mes} de {d.year}')
        start_row, dc, qc, vc, tc, tot = 24,'C','D','E','F','F57'

    varPcts = VARS.get(cid, [0.1] * 35)
    total   = 0.0

    for i in range(35):
        row = start_row + i
        it  = itens[i] if i < len(itens) else None
        pct = varPcts[i % len(varPcts)]
        if it:
            orig = it.get('produto') or it.get('descricao', '')
            desc = orig
            if vars_ia and vars_ia.get('variacoes'):
                match = next((x for x in vars_ia['variacoes'] if x.get('original') == orig), None)
                if match:
                    desc = match.get(cid, orig)
            vu  = round(parse_float(it.get('valorUnitario', 0)) * (1 + pct), 2)
            qtd = parse_float(it.get('qtd', 1))
            vt  = round(vu * qtd, 2)
            total += vt
            set_cell(ws, f'{dc}{row}', desc)
            set_cell(ws, f'{qc}{row}', qtd)
            set_cell(ws, f'{vc}{row}', vu)
            set_cell(ws, f'{tc}{row}', vt)
        else:
            set_cell(ws, f'{dc}{row}', None)
            set_cell(ws, f'{qc}{row}', None)
            set_cell(ws, f'{vc}{row}', None)
            set_cell(ws, f'{tc}{row}', None)

    set_cell(ws, tot, round(total, 2))
    buf = io.BytesIO()
    wb.save(buf)
    return reinjetar_imagens(buf.getvalue(), path)

def handler(event, context):
    if not HAS_OPENPYXL:
        return {'statusCode': 500, 'body': json.dumps({'error': 'openpyxl nao instalado'})}

    if event.get('httpMethod') != 'POST':
        return {'statusCode': 405, 'body': 'Method Not Allowed'}

    try:
        body       = json.loads(event.get('body', '{}'))
        itens      = body.get('itens', []) or []
        cli        = body.get('cliente')
        onum       = body.get('onum', '1')
        tipo       = body.get('tipo', 'Material')
        chips      = body.get('chips', []) or []
        vars_ia    = body.get('vars')
        odata_str  = body.get('odata')
        datas_conc = body.get('datasConc') or []

        odata = datetime.strptime(odata_str, '%Y-%m-%d') if odata_str else datetime.today()

        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            prefix = 'ORC' if tipo == 'Material' else 'NF'
            zf.writestr(f'{prefix}_EGF_{onum}.xlsx',
                        preencher_egf(itens, cli, odata, onum, tipo))
            for i, cid in enumerate(chips):
                if datas_conc and i < len(datas_conc):
                    dc = datetime.strptime(datas_conc[i], '%Y-%m-%d')
                else:
                    delta = -3 if i == 0 else 2
                    dc = odata + timedelta(days=delta)
                zf.writestr(f'{prefix}_{cid}_{onum}.xlsx',
                            preencher_concorrente(cid, itens, cli, dc, tipo, vars_ia))

        return {
            'statusCode': 200,
            'headers': {'Content-Type': 'application/json'},
            'body': json.dumps({'zip': base64.b64encode(zip_buf.getvalue()).decode()})
        }

    except Exception as e:
        import traceback
        return {
            'statusCode': 500,
            'body': json.dumps({'error': str(e), 'trace': traceback.format_exc()})
        }
